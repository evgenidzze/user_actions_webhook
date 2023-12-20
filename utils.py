import logging
import os
from datetime import datetime
from aiomysql import Pool, cursors, create_pool
import mysql.connector
from google.oauth2.credentials import Credentials
from google.oauth2 import service_account
from openpyxl.utils import get_column_letter
import googleapiclient.discovery

SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SERVICE_ACCOUNT_FILE = 'key.json'
SAMPLE_SPREADSHEET_ID = '1Hjd6Oz18FqEH0ayJdcUuCBLI1HTdPf3_kMLm1bN0n60'
creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)

if os.path.exists("token.json"):
    creds = Credentials.from_authorized_user_file("token.json", SCOPES)
service = googleapiclient.discovery.build("sheets", "v4", credentials=creds)
sheet = service.spreadsheets()


def insert_user_data_excel(users: list, x=None):
    start_column = get_column_letter(x)
    end_column = get_column_letter(x + 3)
    result = sheet.values().get(spreadsheetId=SAMPLE_SPREADSHEET_ID, range=f'TEST!{start_column}:{start_column}',
                                majorDimension="COLUMNS").execute()

    last_row = len(result.get('values')[0]) + 1
    print(f'{start_column}{last_row}:{end_column}{last_row}')
    data_rows = []
    for user_data in users:
        data_rows.append([user_data.get('action_time'), user_data.get('username'), user_data.get('user_id'), user_data.get('state')])
    result = sheet.values().append(
        spreadsheetId=SAMPLE_SPREADSHEET_ID,
        range=f'TEST!{start_column}{last_row}:{end_column}{last_row}',
        body={"values": data_rows},
        valueInputOption="RAW"
    ).execute()


connection = mysql.connector.connect(
    host='130.0.238.226',
    user='Zhenya',
    password='AaLaBu14!W',
    database='Signals_avi',
)
cursor = connection.cursor()


def get_all_users():
    cursor = connection.cursor()
    query = f"SELECT * FROM Signals_avi.user_actions"
    cursor.execute(query)
    result = cursor.fetchall()
    return result



